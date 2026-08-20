"""Export a preview sample of the segment-classifier training dataset.

    python ml/personalization/export_dataset_sample.py

There is no dataset *file* in this repo. The training matrix is materialised at
train time by `ml/personalization/train.py`, which queries the database and runs
`app/ml/features.py` over every labelled user. This script walks the identical
code path -- same query, same `build_features`, same `MIN_EVENTS` filter -- and
writes the result to a spreadsheet so the matrix can be inspected without
running a training job.

It is an exporter, not a generator: it invents nothing, rounds nothing, and
derives no column the model does not receive. If the numbers here disagree with
`model/training_report.json`, the database has been reseeded since the shipped
model was fitted, and the sheet says so.
"""

from __future__ import annotations

import random
import sys
from collections import Counter
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO / "backend"))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.db.database import SessionLocal  # noqa: E402
from app.ml.features import FEATURE_NAMES, build_features  # noqa: E402
from app.models.user import UserPreference  # noqa: E402

# Mirrors train.py. Kept as a literal import target rather than a copy so the
# two cannot drift apart silently.
from train import MIN_EVENTS  # noqa: E402  (same directory)

OUT_PATH = Path(__file__).resolve().parent / "training_dataset_sample.xlsx"

#: Rows per segment in the preview sheet. Stratified so all four classes show.
PER_SEGMENT = 10
SAMPLE_SEED = 42

FONT = "Arial"

# What each column means, in FEATURE_NAMES order. Sourced from app/ml/features.py.
DESCRIPTIONS: dict[str, tuple[str, str]] = {
    "n_events": ("count", "Total recorded interactions for this user. Rows below the MIN_EVENTS floor are dropped before training."),
    "rate_viewed": ("fraction 0-1", "Share of this user's events that were 'viewed'."),
    "rate_clicked": ("fraction 0-1", "Share of this user's events that were 'clicked'."),
    "rate_liked": ("fraction 0-1", "Share of this user's events that were 'liked'."),
    "rate_saved": ("fraction 0-1", "Share of this user's events that were 'saved'."),
    "rate_purchased": ("fraction 0-1", "Share of this user's events that were 'purchased'."),
    "rate_disliked": ("fraction 0-1", "Share of this user's events that were 'disliked'."),
    "rate_not_interested": ("fraction 0-1", "Share of this user's events that were 'not_interested'."),
    "engagement_rate": ("fraction 0-1", "(clicked + liked + saved) / n_events."),
    "conversion_rate": ("fraction 0-1", "purchased / n_events."),
    "negative_rate": ("fraction 0-1", "(disliked + not_interested) / n_events."),
    "mean_price": ("INR", "Mean list price of every product the user touched."),
    "median_price": ("INR", "Median list price of every product the user touched."),
    "price_spread": ("INR", "Sample standard deviation of those prices."),
    "mean_discount": ("percentage points", "Mean discount_pct across touched products. Catalog range is 0-48."),
    "max_discount": ("percentage points", "Largest discount_pct the user was exposed to."),
    "mean_rating": ("stars 0-5", "Mean catalog rating of touched products."),
    "mean_review_count": ("count", "Mean review count of touched products."),
    "mean_delivery_days": ("days", "Mean advertised delivery time of touched products."),
    "n_distinct_products": ("count", "Distinct products touched."),
    "n_distinct_categories": ("count", "Distinct categories touched."),
    "n_distinct_brands": ("count", "Distinct brands touched."),
    "brand_concentration": ("fraction 0-1", "Share of events on the single most-frequent brand."),
    "category_concentration": ("fraction 0-1", "Share of events on the single most-frequent category."),
    "purchase_mean_discount": ("percentage points", "Mean discount_pct on purchased items only. 0 when the user bought nothing."),
    "purchase_mean_price": ("INR", "Mean price of purchased items only. 0 when the user bought nothing."),
    "discount_gap": ("percentage points", "purchase_mean_discount minus mean discount on viewed/clicked events. Positive means discounts convert this user."),
}

RUPEE = '"₹"#,##0'
FRACTION = "0.0000"
DECIMAL = "0.00"
INTEGER = "#,##0"

FORMATS: dict[str, str] = {
    "n_events": INTEGER,
    "mean_price": RUPEE,
    "median_price": RUPEE,
    "price_spread": RUPEE,
    "purchase_mean_price": RUPEE,
    "mean_discount": DECIMAL,
    "max_discount": DECIMAL,
    "purchase_mean_discount": DECIMAL,
    "discount_gap": DECIMAL,
    "mean_rating": DECIMAL,
    "mean_review_count": INTEGER,
    "mean_delivery_days": DECIMAL,
    "n_distinct_products": INTEGER,
    "n_distinct_categories": INTEGER,
    "n_distinct_brands": INTEGER,
}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
LABEL_FILL = PatternFill("solid", fgColor="EDE9FB")
NOTE_FILL = PatternFill("solid", fgColor="FFF7E6")
THIN = Side(style="thin", color="D0CCD8")


def load_rows():
    """The exact `train.py::load_dataset` path, with identifiers kept."""
    with SessionLocal() as db:
        labels = {
            row.user_id: row.segment
            for row in db.execute(select(UserPreference.user_id, UserPreference.segment))
            if row.segment
        }
        features = build_features(db, list(labels))

    rows = [f for f in features if f.n_events >= MIN_EVENTS and labels.get(f.user_id)]
    if not rows:
        raise SystemExit(
            "No labelled users with enough history.\n"
            "Seed them first:  python -m scripts.seed --interactions 15000"
        )
    rows.sort(key=lambda r: r.user_id)
    return rows, labels


def stratified_sample(rows, labels):
    by_segment: dict[str, list] = {}
    for row in rows:
        by_segment.setdefault(labels[row.user_id], []).append(row)

    rng = random.Random(SAMPLE_SEED)
    picked = []
    for segment in sorted(by_segment):
        pool = by_segment[segment]
        picked.extend(rng.sample(pool, min(PER_SEGMENT, len(pool))))
    picked.sort(key=lambda r: (labels[r.user_id], r.user_id))
    return picked


def style_header(ws, row_idx: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)


def write_readme(wb, rows, labels, report_matches: bool, report: dict) -> None:
    ws = wb.create_sheet("README", 0)
    ws.sheet_properties.tabColor = "4B35C9"

    lines = [
        ("SmartBuy AI - shopper-segment classifier: training dataset preview", True),
        ("", False),
        ("WHAT THIS FILE IS", True),
        ("A sample of the exact matrix that ml/personalization/train.py fits the model on.", False),
        ("There is no dataset file in the repository. The matrix is built at train time from", False),
        ("the application database by app/ml/features.py, which is the same feature code that", False),
        ("runs at serving time. This export walks that identical path and writes the result out.", False),
        ("", False),
        ("HOW A ROW IS PRODUCED", True),
        ("1. Every user in user_preferences that carries a non-null `segment` is selected.", False),
        ("2. app/ml/features.py aggregates that user's rows in product_interactions, joined to", False),
        ("   products, into the 27 numbers in FEATURE_NAMES order.", False),
        (f"3. Users with fewer than {MIN_EVENTS} interactions are dropped (MIN_EVENTS in train.py).", False),
        ("4. `segment` is the label. It is never an input feature.", False),
        ("", False),
        ("WHAT THE DATA IS", True),
        ("Synthetic. Every user, interaction and segment label in this matrix was generated by", False),
        ("backend/scripts/seed.py. No real shopper data was collected or used, and there is no", False),
        ("PII in this file -- the identifiers are generated UUIDs.", False),
        ("", False),
        ("The catalog the interactions point at is a curated product set; prices, ratings and", False),
        ("discounts are catalog attributes, not live marketplace prices.", False),
        ("", False),
        ("THE CAVEAT THAT MUST TRAVEL WITH THE ACCURACY FIGURE", True),
        ("Because seed.py both assigns the segment and shapes the behaviour that follows from it,", False),
        ("a model that recovers the segment demonstrates the pipeline works end to end. It is not", False),
        ("evidence about how real shoppers behave. The holdout accuracy in training_report.json", False),
        ("must never be presented as a claim about real customers.", False),
        ("", False),
        ("SHEETS", True),
        ("training_sample    Stratified preview rows, verbatim feature values.", False),
        ("feature_schema     All 27 columns: order, unit, and definition.", False),
        ("class_balance      Label distribution, in the sample and in the full matrix.", False),
        ("", False),
        ("REPRODUCE", True),
        ("python ml/personalization/export_dataset_sample.py", False),
        (f"Sampling is deterministic: {PER_SEGMENT} rows per segment, random.Random({SAMPLE_SEED}).", False),
    ]

    for i, (text, is_heading) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name=FONT, bold=is_heading, size=11 if i == 1 else 10)
        cell.alignment = Alignment(vertical="top")

    row = len(lines) + 2
    ws.cell(row=row, column=1, value="INTEGRITY CHECK AGAINST THE SHIPPED MODEL").font = Font(
        name=FONT, bold=True, size=10
    )
    row += 1
    if report_matches:
        verdict = (
            f"Match. This export rebuilt {len(rows)} rows x {len(FEATURE_NAMES)} features with class "
            f"counts identical to model/training_report.json, so the database still holds the "
            f"dataset the shipped classifier was fitted on."
        )
    else:
        verdict = (
            f"MISMATCH. This export rebuilt {len(rows)} rows x {len(FEATURE_NAMES)} features with "
            f"class counts {dict(Counter(labels[r.user_id] for r in rows))}, but "
            f"model/training_report.json records {report.get('n_users')} rows and "
            f"{report.get('class_counts')}. The database has been reseeded since the shipped model "
            f"was trained -- these rows are the current matrix, not the one that produced the "
            f"saved classifier."
        )
    cell = ws.cell(row=row, column=1, value=verdict)
    cell.font = Font(name=FONT, size=10)
    cell.fill = NOTE_FILL
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 46
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    ws.column_dimensions["A"].width = 100
    ws.sheet_view.showGridLines = False


def write_sample(wb, sample, labels) -> str:
    ws = wb.create_sheet("training_sample")
    headers = ["user_id", "segment (LABEL)", *FEATURE_NAMES]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for row in sample:
        ws.append([row.user_id, labels[row.user_id], *row.values])

    for r in range(2, len(sample) + 2):
        ws.cell(row=r, column=1).font = Font(name=FONT, size=9, color="6B6478")
        label_cell = ws.cell(row=r, column=2)
        label_cell.font = Font(name=FONT, size=10, bold=True)
        label_cell.fill = LABEL_FILL
        for i, name in enumerate(FEATURE_NAMES):
            cell = ws.cell(row=r, column=3 + i)
            cell.font = Font(name=FONT, size=10)
            cell.number_format = FORMATS.get(name, FRACTION)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 17
    for i in range(len(FEATURE_NAMES)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 34

    note_row = len(sample) + 3
    note = ws.cell(
        row=note_row,
        column=1,
        value=(
            "Values are written verbatim from app/ml/features.py -- nothing on this sheet is "
            "rounded, scaled or recomputed. Cell display formats are cosmetic only; the stored "
            "value is the exact float the model receives. 'segment' is the target, not an input."
        ),
    )
    note.font = Font(name=FONT, size=9, italic=True)
    note.fill = NOTE_FILL
    note.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    ws.row_dimensions[note_row].height = 30
    return ws.title


def write_schema(wb) -> None:
    ws = wb.create_sheet("feature_schema")
    ws.append(["#", "column", "unit", "definition"])
    style_header(ws, 1, 4)

    for i, name in enumerate(FEATURE_NAMES):
        unit, definition = DESCRIPTIONS[name]
        ws.append([i, name, unit, definition])

    for r in range(2, len(FEATURE_NAMES) + 2):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10, bold=(c == 2))
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 4))

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 19
    ws.column_dimensions["D"].width = 88
    ws.freeze_panes = "A2"

    row = len(FEATURE_NAMES) + 3
    for text in (
        "Column order is FEATURE_NAMES in app/ml/features.py. The order is the contract: the",
        "model is persisted alongside this list, and serving refuses to predict if the two differ.",
        "The label column (segment) is not in this list and is never fed to the model.",
    ):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name=FONT, size=9, italic=True)
        row += 1


def write_class_balance(wb, sample_sheet: str, sample, rows, labels) -> None:
    ws = wb.create_sheet("class_balance")
    full_counts = Counter(labels[r.user_id] for r in rows)
    first = 2
    last = len(sample) + 1

    ws.append(["segment", "rows in this sample", "rows in full matrix", "share of full matrix"])
    style_header(ws, 1, 4)

    for i, segment in enumerate(sorted(full_counts)):
        r = 2 + i
        ws.cell(row=r, column=1, value=segment)
        # Counted from the sample sheet rather than from Python, so the figure
        # recalculates if a row is edited or removed.
        ws.cell(
            row=r,
            column=2,
            value=f"=COUNTIF('{sample_sheet}'!$B${first}:$B${last},$A{r})",
        )
        ws.cell(row=r, column=3, value=int(full_counts[segment]))
        ws.cell(row=r, column=4, value=f"=IF($C${total_row(full_counts)}=0,0,C{r}/$C${total_row(full_counts)})")

    total = total_row(full_counts)
    ws.cell(row=total, column=1, value="TOTAL")
    ws.cell(row=total, column=2, value=f"=SUM(B2:B{total - 1})")
    ws.cell(row=total, column=3, value=f"=SUM(C2:C{total - 1})")
    ws.cell(row=total, column=4, value=f"=SUM(D2:D{total - 1})")

    for r in range(2, total + 1):
        bold = r == total
        ws.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=bold)
        for c in (2, 3):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10, bold=bold)
            cell.number_format = INTEGER
        pct = ws.cell(row=r, column=4)
        pct.font = Font(name=FONT, size=10, bold=bold)
        pct.number_format = "0.0%"
        if bold:
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = Border(top=THIN)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    note = ws.cell(
        row=total + 2,
        column=1,
        value=(
            "Column B counts the preview sheet live. Column C is the full matrix this export "
            "rebuilt from the database; train.py splits it 75/25 stratified, so both halves keep "
            "these proportions."
        ),
    )
    note.font = Font(name=FONT, size=9, italic=True)
    note.alignment = Alignment(vertical="top", wrap_text=True)
    ws.merge_cells(start_row=total + 2, start_column=1, end_row=total + 2, end_column=4)
    ws.row_dimensions[total + 2].height = 30


def total_row(full_counts: Counter) -> int:
    return len(full_counts) + 2


def main() -> int:
    import json

    rows, labels = load_rows()
    sample = stratified_sample(rows, labels)

    report_path = Path(__file__).resolve().parent / "model" / "training_report.json"
    report = json.loads(report_path.read_text(encoding="utf-8")) if report_path.exists() else {}
    counts = {k: int(v) for k, v in Counter(labels[r.user_id] for r in rows).items()}
    report_matches = (
        report.get("n_users") == len(rows)
        and report.get("n_features") == len(FEATURE_NAMES)
        and report.get("class_counts") == counts
    )

    wb = Workbook()
    wb.remove(wb.active)
    write_readme(wb, rows, labels, report_matches, report)
    sample_sheet = write_sample(wb, sample, labels)
    write_schema(wb)
    write_class_balance(wb, sample_sheet, sample, rows, labels)

    wb.save(OUT_PATH)
    print(f"full matrix   {len(rows)} rows x {len(FEATURE_NAMES)} features")
    print(f"classes       {counts}")
    print(f"sample        {len(sample)} rows ({PER_SEGMENT} per segment, seed {SAMPLE_SEED})")
    print(f"report match  {report_matches}")
    print(f"saved         {OUT_PATH.relative_to(REPO)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
